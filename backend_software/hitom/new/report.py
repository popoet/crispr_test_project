import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Font


def find_gap_regions(reference_sequence):
    """
    统计 reference_sequence 中连续 '-' 的区间。
    """
    gap_regions = []
    in_gap = False
    start = 0

    for i, char in enumerate(reference_sequence):
        if char == '-' and not in_gap:
            in_gap = True
            start = i
        elif char != '-' and in_gap:
            in_gap = False
            gap_regions.append((start, i - 1))

    if in_gap:
        gap_regions.append((start, len(reference_sequence) - 1))

    return gap_regions


def adjust_range(reference_sequence, start, end):
    if start < 0 or start > end:
        raise ValueError("Invalid start or end positions.")

    gap_regions = find_gap_regions(reference_sequence)
    adjusted_start = start
    adjusted_end = end

    for gap_start, gap_end in gap_regions:
        if gap_end <= start:
            adjusted_start += (gap_end - gap_start + 1)
            adjusted_end += (gap_end - gap_start + 1)
        elif gap_start >= start and gap_end <= end:
            adjusted_end += (gap_end - gap_start + 1)
        elif gap_start <= start and gap_end >= end:
            total_gaps = sum(gap_end - gap_start + 1 for gap_start, gap_end in gap_regions)
            return start + total_gaps, end + total_gaps
        elif gap_start > end:
            break
        elif gap_start <= end < gap_end:
            non_gap_bases = 0
            for i in range(start, len(reference_sequence)):
                if reference_sequence[i] != '-':
                    non_gap_bases += 1
                if non_gap_bases >= (end - start + 1):
                    adjusted_end = i
                    break

    return adjusted_start, adjusted_end


def merge_variants(variant_types, variant_seqs, variant_positions):
    merged_types = []
    merged_seqs = []
    merged_positions = []

    i = 0
    while i < len(variant_types):
        current_type = variant_types[i]
        current_pos = variant_positions[i]
        current_seq = variant_seqs[i]

        if current_type == "SNP":
            start_pos = current_pos
            end_pos = current_pos
            snp_ref = current_seq[0]
            snp_alt = current_seq[-1]
            j = i + 1
            while j < len(variant_types) and variant_types[j] == "SNP" and variant_positions[j] == end_pos + 1:
                end_pos = variant_positions[j]
                snp_ref += variant_seqs[j][0]
                snp_alt += variant_seqs[j][-1]
                j += 1
            if start_pos == end_pos:
                merged_types.append("SNP")
                merged_seqs.append(current_seq)
                merged_positions.append(str(start_pos))
            else:
                merged_types.append("SNP")
                merged_seqs.append(f"{snp_ref}->{snp_alt}")
                merged_positions.append(f"{start_pos}-{end_pos}")
            i = j
        elif current_type.endswith("I") or current_type.endswith("D"):
            length = int(current_type[:-1])
            start_pos = current_pos
            end_pos = current_pos
            combined_seq = current_seq
            j = i + 1
            while j < len(variant_types) and variant_types[j] == current_type and variant_positions[j] == end_pos + 1:
                end_pos = variant_positions[j]
                combined_seq += variant_seqs[j]
                length += int(variant_types[j][:-1])
                j += 1
            merged_types.append(f"{length}{current_type[-1]}")
            merged_seqs.append(combined_seq)
            if start_pos == end_pos:
                merged_positions.append(str(start_pos))
            else:
                merged_positions.append(f"{start_pos}-{end_pos}")
            i = j
        else:
            merged_types.append(current_type)
            merged_seqs.append(current_seq)
            merged_positions.append(str(current_pos))
            i += 1

    if merged_types.count("SNP") > 1:
        merged_types = ["SNP"] + [t for t in merged_types if t != "SNP"]

    return merged_types, merged_seqs, merged_positions


def extract_variants_with_merge(input_file, output_file, start, end):
    # 计算总的 #Reads 数
    total_reads = 0
    rows = []
    with open(input_file, 'r') as infile:
        reader = csv.DictReader(infile, delimiter='\t')
        for row in reader:
            rows.append(row)
            total_reads += int(row['#Reads'])

    # 存储变异信息
    variant_data = []

    for row in rows:
        aligned_seq = row['Aligned_Sequence']
        reference_seq = row['Reference_Sequence']
        num_reads = int(row['#Reads'])
        perc_reads = num_reads / total_reads * 100

        adjusted_start, adjusted_end = adjust_range(reference_seq, start, end)
        aligned_subseq = aligned_seq[adjusted_start:adjusted_end]
        reference_subseq = reference_seq[adjusted_start:adjusted_end]

        variant_types = []
        variant_seqs = []
        variant_positions = []
        amp_seq = aligned_subseq.upper()

        for i, (ref, align) in enumerate(zip(reference_subseq, aligned_subseq)):
            actual_position = adjusted_start + i
            if ref != align:
                if ref == "-":
                    variant_types.append("1I")
                    variant_seqs.append(align)
                    variant_positions.append(actual_position)
                    amp_seq = amp_seq[:i] + align.lower() + amp_seq[i + 1:]
                elif align == "-":
                    variant_types.append("1D")
                    variant_seqs.append(ref)
                    variant_positions.append(actual_position)
                    amp_seq = amp_seq[:i] + '-' + amp_seq[i + 1:]
                else:
                    variant_types.append("SNP")
                    variant_seqs.append(f"{ref}->{align}")
                    variant_positions.append(actual_position)
                    amp_seq = amp_seq[:i] + align.lower() + amp_seq[i + 1:]

        merged_types, merged_seqs, merged_positions = merge_variants(
            variant_types, variant_seqs, variant_positions
        )

        if not merged_types:
            merged_types = ["WT"]
            merged_seqs = ["-"]
            merged_positions = ["未检测到变异"]

        perc_reads_display = f"{perc_reads:.2f}%" if perc_reads > 0 else "<0.2%"
        variant_type_str = ";".join(merged_types)
        variant_seq_str = ";".join(merged_seqs)
        variant_positions_str = ";".join(map(str, merged_positions))

        variant_data.append({
            '变异编号': len(variant_data) + 1,
            '支持reads数目': num_reads,
            '变异比例': perc_reads_display,
            '变异类型': variant_type_str,
            '变异序列': variant_seq_str,
            '扩增片段序列': amp_seq,
            '变异位置': variant_positions_str
        })

    # 合并相同扩增片段的变异
    merged_variant_data = {}
    for variant in variant_data:
        amp_seq = variant['扩增片段序列']
        if amp_seq not in merged_variant_data:
            merged_variant_data[amp_seq] = {
                '支持reads数目': 0,
                '变异比例': 0.0,
                '变异类型': set(),
                '变异序列': set(),
                '变异位置': set(),
                '变异编号': variant['变异编号']
            }

        merged_variant_data[amp_seq]['支持reads数目'] += int(variant['支持reads数目'])
        merged_variant_data[amp_seq]['变异比例'] += float(variant['变异比例'].replace('%', ''))
        merged_variant_data[amp_seq]['变异类型'].update(variant['变异类型'].split(';'))
        merged_variant_data[amp_seq]['变异序列'].update(variant['变异序列'].split(';'))
        merged_variant_data[amp_seq]['变异位置'].update(variant['变异位置'].split(';'))

    # 对结果排序
    sorted_variants = sorted(
        merged_variant_data.items(),
        key=lambda item: item[1]['支持reads数目'],
        reverse=True
    )

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "变异分析结果"

    # 设置表头样式
    header_font = Font(bold=True)
    headers = ['变异编号', '支持reads数目', '变异比例', '变异类型', '变异序列', '扩增片段序列']
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font

    # 写入数据
    new_variant_index = 1
    for row_num, (amp_seq, data) in enumerate(sorted_variants, 2):
        ws.cell(row=row_num, column=1, value=new_variant_index)
        ws.cell(row=row_num, column=2, value=data['支持reads数目'])
        ws.cell(row=row_num, column=3, value=f"{data['变异比例']:.2f}%" if data['变异比例'] > 0 else "<0.2%")
        ws.cell(row=row_num, column=4, value=";".join(data['变异类型']))
        ws.cell(row=row_num, column=5, value=";".join(data['变异序列']))
        ws.cell(row=row_num, column=6, value=amp_seq)
        new_variant_index += 1

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存Excel文件
    if not output_file.endswith('.xlsx'):
        output_file = output_file + '.xlsx'
    wb.save(output_file)
    print(f"提取完成，变异结果已保存至 {output_file}")


if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("用法: python report.py <input_file> <start> <end> <output_file>")
        sys.exit(1)

    input_file = sys.argv[1]
    start = int(sys.argv[2])
    end = int(sys.argv[3])
    output_file = sys.argv[4]

    extract_variants_with_merge(input_file, output_file, start, end)
